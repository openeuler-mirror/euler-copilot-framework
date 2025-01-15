"""用户导出工具

Copyright (c) Huawei Technologies Co., Ltd. 2023-2024. All rights reserved.
"""
from __future__ import annotations

import argparse
import datetime
import re
import secrets
import shutil
import zipfile
from pathlib import Path
from typing import ClassVar

from openpyxl import Workbook

from apps.common.security import Security
from apps.entities.collection import Audit
from apps.manager.audit_log import AuditLogManager
from apps.manager.conversation import ConversationManager
from apps.manager.record import RecordManager
from apps.manager.user import UserManager


class UserExporter:
    """用户导出工具类"""

    start_row_id: int = 1
    chat_xlsx_column: ClassVar[list[str]] = ["question", "answer", "created_time"]
    chat_column_map: ClassVar[dict[str, int]] = {
        "question_column": 1,
        "answer_column": 2,
        "created_time_column": 3,
    }
    user_info_xlsx_column: ClassVar[list[str]] = [
        "user_sub", "organization",
        "created_time", "login_time", "revision_number",
    ]
    user_info_column_map: ClassVar[dict[str, int]] = {
        "user_sub_column": 1,
        "organization_column": 2,
        "created_time_column": 3,
        "login_time_column": 4,
        "revision_number_column": 5,
    }

    @staticmethod
    def get_datetime_from_str(date_str: str, date_format: str) -> float:
        """将日期字符串转换为时间戳"""
        date_time_obj = datetime.datetime.strptime(date_str, date_format).astimezone(datetime.timezone.utc)
        date_time_obj = datetime.datetime(date_time_obj.year, date_time_obj.month, date_time_obj.day).astimezone(datetime.timezone.utc)
        return date_time_obj.timestamp()

    @staticmethod
    def zip_xlsx_folder(tmp_out_dir: Path) -> Path:
        """将xlsx文件夹压缩为zip文件"""
        dir_name = tmp_out_dir.parent
        last_dir_name = tmp_out_dir.name
        xlsx_file_name_list = list(tmp_out_dir.glob("*.xlsx"))
        zip_file_dir = Path(dir_name) / (last_dir_name + ".zip")
        with zipfile.ZipFile(zip_file_dir, "w") as zip_file:
            for xlsx_file_name in xlsx_file_name_list:
                xlsx_file_path = tmp_out_dir / xlsx_file_name
                zip_file.write(xlsx_file_path)
        return zip_file_dir

    @staticmethod
    def save_chat_to_xlsx(xlsx_dir, chat_list):
        """将聊天记录保存到xlsx文件中"""
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            err = "Workbook没有active的sheet"
            raise ValueError(err)
        for i, column in enumerate(UserExporter.chat_xlsx_column):
            sheet.cell(row=UserExporter.start_row_id, column=i+1, value=column)
        row_id = UserExporter.start_row_id + 1
        for chat in chat_list:
            question = chat[0]
            answer = chat[1]
            created_time = chat[2]
            sheet.cell(row=row_id,
                       column=UserExporter.chat_column_map["question_column"],
                       value=question)
            sheet.cell(row=row_id,
                       column=UserExporter.chat_column_map["answer_column"],
                       value=answer)
            sheet.cell(row=row_id,
                       column=UserExporter.chat_column_map["created_time_column"],
                       value=created_time)
            row_id += 1
        workbook.save(xlsx_dir)

    @staticmethod
    def save_user_info_to_xlsx(xlsx_dir, user_info):
        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            err = "Workbook没有active的sheet"
            raise ValueError(err)
        for i, column in enumerate(UserExporter.user_info_xlsx_column):
            sheet.cell(row=UserExporter.start_row_id, column=i+1, value=column)
        row_id = UserExporter.start_row_id + 1
        user_sub = user_info.user_sub
        organization = user_info.organization
        created_time = user_info.created_time
        login_time = user_info.login_time
        revision_number = user_info.revision_number
        sheet.cell(row=row_id,
                   column=UserExporter.user_info_column_map["user_sub_column"],
                   value=user_sub)
        sheet.cell(row=row_id,
                   column=UserExporter.user_info_column_map["organization_column"],
                   value=organization)
        sheet.cell(row=row_id,
                   column=UserExporter.user_info_column_map["created_time_column"],
                   value=created_time)
        sheet.cell(row=row_id,
                   column=UserExporter.user_info_column_map["login_time_column"],
                   value=login_time)
        sheet.cell(row=row_id,
                   column=UserExporter.user_info_column_map["revision_number_column"],
                   value=revision_number)
        workbook.save(xlsx_dir)

    @staticmethod
    def export_user_info_to_xlsx(tmp_out_dir, user_sub):
        user_info = UserManager.get_userinfo_by_user_sub(user_sub)
        xlsx_file_name = "user_info_" + user_sub + ".xlsx"
        xlsx_file_name = re.sub(r'[<>:"/\\|?*]', "_", xlsx_file_name)
        xlsx_file_name = xlsx_file_name.replace(" ", "_")
        xlsx_dir = Path(tmp_out_dir) / xlsx_file_name
        UserExporter.save_user_info_to_xlsx(xlsx_dir, user_info)

    @staticmethod
    def export_chats_to_xlsx(tmp_out_dir, user_sub, start_day, end_day):
        user_qa_records = ConversationManager.get_conversation_by_user_sub(
            user_sub)
        for user_qa_record in user_qa_records:
            chat_id = user_qa_record.conversation_id
            chat_tile = re.sub(r'[<>:"/\\|?*]', "_", user_qa_record.title)
            chat_tile = chat_tile.replace(" ", "_")[:20]
            chat_created_time = str(user_qa_record.created_time)
            encrypted_qa_records = RecordManager.query_encrypted_data_by_conversation_id(
                chat_id)
            chat = []
            for record in encrypted_qa_records:
                question = Security.decrypt(record.encrypted_question,
                                            record.question_encryption_config)
                answer = Security.decrypt(record.encrypted_answer,
                                          record.answer_encryption_config)
                qa_record_created_time = record.created_time
                if start_day is not None and UserExporter.get_datetime_from_str(record.created_time, "%Y-%m-%d %H:%M:%S") < start_day:
                    continue
                if end_day is not None and UserExporter.get_datetime_from_str(record.created_time, "%Y-%m-%d %H:%M:%S") > end_day:
                    continue
                chat.append([question, answer, qa_record_created_time])
            xlsx_file_name = "chat_"+chat_tile[:20] + "_"+chat_created_time+".xlsx"
            xlsx_file_name = xlsx_file_name.replace(" ", "")
            xlsx_dir = Path(tmp_out_dir) / xlsx_file_name
            UserExporter.save_chat_to_xlsx(xlsx_dir, chat)

    @staticmethod
    def export_user_data(users_dir, user_sub, export_preferences=None, start_day=None, end_day=None):
        export_preferences = export_preferences or ["user_info", "chat"]
        rand_num = secrets.randbits(128)
        tmp_out_dir = Path("./") / users_dir / str(rand_num)
        if tmp_out_dir.exists():
            shutil.rmtree(tmp_out_dir)
        tmp_out_dir.mkdir(parents=True, exist_ok=True)
        tmp_out_dir.chmod(0o750)
        if "user_info" in export_preferences:
            UserExporter.export_user_info_to_xlsx(tmp_out_dir, user_sub)
        if "chat" in export_preferences:
            UserExporter.export_chats_to_xlsx(tmp_out_dir, user_sub, start_day, end_day)
        zip_file_path = UserExporter.zip_xlsx_folder(tmp_out_dir)
        shutil.rmtree(tmp_out_dir)
        return zip_file_path


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--user_sub", type=str, required=True,
                        help="""Please provide usr_sub identifier for the export \
                        process. This ID ensures that the exported data is \
                        accurately associated with your user profile. If this \
                        field is \"all\", then all user information will be  \
                        exported""")
    parser.add_argument("--export_preferences", type=str, required=True,
                        help="""Please enter your export preferences by specifying \
                        'chat' and/or 'user_info', separated by a space \
                        if including both. Ensure that your input is limited to \
                        these options for accurate data export processing.""")
    parser.add_argument("--start_day", type=str, required=False,
                        help="""User record export start date, format reference is \
                        as follows: 2024_03_23""")
    parser.add_argument("--end_day", type=str, required=False,
                        help="""User record export end date, format reference is \
                        as follows: 2024_03_23""")
    args = vars(parser.parse_args())
    arg_user_sub = args["user_sub"]
    arg_export_preferences = args["export_preferences"].split(" ")
    start_day = args["start_day"]
    end_day = args["end_day"]
    try:
        if start_day is not None:
            start_day = UserExporter.get_datetime_from_str(start_day, "%Y_%m_%d")
    except Exception as e:
        data = Audit(
            user_sub=arg_user_sub,
            http_method="internal_user_exporter",
            module="export_user_data",
            client_ip="internal",
            message=f"start_day_exchange failed due error: {e}",
        )
        AuditLogManager.add_audit_log(data)
    try:
        if end_day is not None:
            end_day = UserExporter.get_datetime_from_str(end_day, "%Y_%m_%d")
    except Exception as e:
        data = Audit(
            user_sub=arg_user_sub,
            http_method="internal_user_exporter",
            module="export_user_data",
            client_ip="internal",
            message=f"end_day_exchange failed due error: {e}",
        )
        AuditLogManager.add_audit_log(data)
    if arg_user_sub == "all":
        user_sub_list = UserManager.get_all_user_sub()
    else:
        user_sub_list = [arg_user_sub]
    users_dir = str(secrets.randbits(128))
    if Path(users_dir).exists():
        shutil.rmtree(users_dir)
    Path(users_dir).mkdir(parents=True, exist_ok=True)
    Path(users_dir).chmod(0o750)
    for arg_user_sub in user_sub_list:
        arg_user_sub = arg_user_sub[0]
        try:
            export_path = UserExporter.export_user_data(
                users_dir, arg_user_sub, arg_export_preferences, start_day, end_day)
            audit_export_preference = f", preference: {arg_export_preferences}" if arg_export_preferences else ""
            data = Audit(
                user_sub=arg_user_sub,
                http_method="internal_user_exporter",
                module="export_user_data",
                client_ip="internal",
                message=f"exported user data of id: {arg_user_sub}{audit_export_preference}, path: {export_path}",
            )
            AuditLogManager.add_audit_log(data)
        except Exception as e:
            data = Audit(
                user_sub=arg_user_sub,
                http_method="internal_user_exporter",
                module="export_user_data",
                client_ip="internal",
                message=f"用户（id: {arg_user_sub}）请求导出数据失败: {e!s}",
            )
            AuditLogManager.add_audit_log(data)
    zip_file_path = UserExporter.zip_xlsx_folder(Path(users_dir))
    shutil.rmtree(users_dir)
